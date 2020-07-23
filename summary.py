import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, colors, Alignment, PatternFill, Border
from datetime import date, datetime
import json

client_name = sys.argv[2]
job_number = sys.argv[1]

# LOAD WORKBOOKS
wb = load_workbook(filename='database/Templates/JobSumTemplate.xlsx')

# PULL IN JOB JSON FILE
with open('database/'+str(job_number)+'/job.json', 'r') as f:
    job_json = json.load(f)

job = job_json

# SET ACTIVATIONS
printout = wb['PRINTOUT']
areas = wb['Areas Summary']
abm = wb['Areas Summary (By Material)']
loc = wb['Spool Location']
shbs = wb['Shorts by Scope']
pbs = wb['Purchased by Scope']
nmbs = wb['No Material by Scope']
mvo = wb['Spool Missing Valve Only']
spbs = wb['Spools by Scope']
disc = wb['Discrepancies']

print('')
print('Adding spools to summary..')

# COLORS
black = '404040'
gray = 'd9d9d9'
header_blue = '538dd5'
data_blue = 'e7eff9'
header_red = 'd27d7c'
data_red = 'fbf3f3'
header_green = 'abc56d'
data_green = 'f4f8ec'
header_orange = 'f79443'
data_orange = 'fdf0e7'
white = 'ffffff'
discrepancies = 'ffb3bc'

# COLUMNS
mat_area = 'A'
mat_priority = 'B'
mat_spools = 'C'
mat_on_hold = 'D'
mat_workable = 'E'
mat_not_workable = 'F'
mat_weldout = 'G'
mat_weldout_remaining = 'H'
mat_stc = 'I'
mat_delivered = 'J'
mat_delivered_remaining = 'K'
mat_perc_workable = 'L'
mat_perc_weldout = 'M'
mat_perc_delivered = 'N'

# AREAS SUMMARY PRINTOUT

# COLUMNS
mat_area = 'A'
mat_priority = 'B'
mat_spools = 'C'
mat_on_hold = 'D'
mat_workable = 'E'
mat_not_workable = 'F'
mat_weldout = 'G'
mat_weldout_remaining = 'H'
mat_stc = 'I'
mat_delivered = 'J'
mat_delivered_remaining = 'K'
mat_perc_workable = 'L'
mat_perc_weldout = 'M'
mat_perc_delivered = 'N'

printout['A1'] = client_name + ': ' + job_number
printout['C1'] = str(job['total']) + " Spools"
printout['F1'] = str(job['workable']) + " Workable"
printout['I1'] = str(job['issued']) + " Issued"

printout['A31'] = job_number + ' Areas Summary'
for area in job['areas']:
    printout[mat_area+str(printout.max_row+1)] = str(area['area'])
    printout[mat_priority+str(printout.max_row)] = str(area['priority'])
    printout[mat_spools+str(printout.max_row)] = area['spools']
    printout[mat_on_hold+str(printout.max_row)] = area['on_hold']
    printout[mat_workable+str(printout.max_row)] = area['workable']
    printout[mat_not_workable+str(printout.max_row)
             ] = area['spools']-area['workable']
    printout[mat_weldout+str(printout.max_row)] = area['weldout']
    printout[mat_weldout_remaining +
             str(printout.max_row)] = area['spools']-area['weldout']
    printout[mat_stc +
             str(printout.max_row)] = area['stc']
    printout[mat_delivered+str(printout.max_row)] = area['delivered']
    printout[mat_delivered_remaining +
             str(printout.max_row)] = area['spools']-area['delivered']
    printout[mat_perc_workable+str(printout.max_row)
             ] = str(area['workable_perc']) + '%'
    printout[mat_perc_weldout+str(printout.max_row)
             ] = str(area['weldout_perc']) + '%'
    printout[mat_perc_delivered+str(printout.max_row)
             ] = str(area['delivered_perc']) + '%'

# TOTALS
printout['A'+str(printout.max_row+1)] = 'TOTALS'
printout[mat_spools+str(printout.max_row)] = job['total']
printout[mat_on_hold+str(printout.max_row)] = job['on_hold']
printout[mat_workable+str(printout.max_row)] = job['workable']
printout[mat_not_workable+str(printout.max_row)
         ] = job['total']-job['workable']
printout[mat_weldout+str(printout.max_row)] = job['weldout']
printout[mat_weldout_remaining +
         str(printout.max_row)] = job['total']-job['weldout']
printout[mat_stc+str(printout.max_row)] = job['stc']
printout[mat_delivered+str(printout.max_row)] = job['delivered']
printout[mat_delivered_remaining +
         str(printout.max_row)] = job['total']-job['delivered']
printout[mat_perc_workable+str(printout.max_row)
         ] = str(round((job['workable']/job['total'])*100))+"%"
printout[mat_perc_weldout+str(printout.max_row)
         ] = str(round((job['weldout']/job['total'])*100))+"%"
printout[mat_perc_delivered+str(printout.max_row)
         ] = str(round((job['delivered']/job['total'])*100))+"%"
printout.row_dimensions[printout.max_row].height = 20

# FILL TOTALS ROW
for col in range(1, printout.max_column+1):
    printout[get_column_letter(col)+str(printout.max_row)
             ].font = Font(bold=True)
    printout[get_column_letter(col)+str(printout.max_row)].fill = PatternFill(
        fill_type='solid', start_color=gray, end_color=gray)

# AREAS SUMMARY TAB
areas['A1'] = 'Areas Summary'
areas[mat_area+'2'] = 'Area'
areas[mat_priority+'2'] = 'Priority'
areas[mat_spools+'2'] = 'Total Spools'
areas[mat_on_hold+'2'] = 'On Hold'
areas[mat_workable+'2'] = 'Workable'
areas[mat_not_workable+'2'] = 'Not Workable'
areas[mat_weldout+'2'] = 'Welded Out'
areas[mat_weldout_remaining+'2'] = 'Remaining to Weld Out'
areas[mat_stc+'2'] = 'Shipped To Paint'
areas[mat_delivered+'2'] = 'Delivered'
areas[mat_delivered_remaining+'2'] = 'Remaining to Deliver'
areas[mat_perc_workable+'2'] = 'Workable %'
areas[mat_perc_weldout+'2'] = 'Weld Out %'
areas[mat_perc_delivered+'2'] = 'Delivered %'

for area in job['areas']:
    areas[mat_area+str(areas.max_row+1)] = area['area']
    areas[mat_priority+str(areas.max_row)] = area['priority']
    areas[mat_spools+str(areas.max_row)] = area['spools']
    areas[mat_on_hold+str(areas.max_row)] = area['on_hold']
    areas[mat_workable+str(areas.max_row)] = area['workable']
    areas[mat_not_workable+str(areas.max_row)
          ] = area['spools']-area['workable']
    areas[mat_weldout+str(areas.max_row)] = area['weldout']
    areas[mat_weldout_remaining +
          str(areas.max_row)] = area['spools']-area['weldout']
    areas[mat_stc +
          str(areas.max_row)] = area['stc']
    areas[mat_delivered+str(areas.max_row)] = area['delivered']
    areas[mat_delivered_remaining +
          str(areas.max_row)] = area['spools']-area['delivered']
    areas[mat_perc_workable+str(areas.max_row)
          ] = str(area['workable_perc']) + '%'
    areas[mat_perc_weldout+str(areas.max_row)
          ] = str(area['weldout_perc']) + '%'
    areas[mat_perc_delivered+str(areas.max_row)
          ] = str(area['delivered_perc']) + '%'

# TOTALS
areas['A'+str(areas.max_row+1)] = 'TOTALS'
areas[mat_spools+str(areas.max_row)] = job['total']
areas[mat_on_hold+str(areas.max_row)] = job['on_hold']
areas[mat_workable+str(areas.max_row)] = job['workable']
areas[mat_not_workable+str(areas.max_row)
      ] = job['total']-job['workable']
areas[mat_weldout+str(areas.max_row)] = job['weldout']
areas[mat_weldout_remaining +
      str(areas.max_row)] = job['total']-job['weldout']
areas[mat_stc+str(areas.max_row)] = job['stc']
areas[mat_delivered+str(areas.max_row)] = job['delivered']
areas[mat_delivered_remaining +
      str(areas.max_row)] = job['total']-job['delivered']
areas[mat_perc_workable+str(areas.max_row)
      ] = str(round((job['workable']/job['total'])*100))+"%"
areas[mat_perc_weldout+str(areas.max_row)
      ] = str(round((job['weldout']/job['total'])*100))+"%"
areas[mat_perc_delivered+str(areas.max_row)
      ] = str(round((job['delivered']/job['total'])*100))+"%"
areas.row_dimensions[areas.max_row].height = 20

# FILL TOTALS ROW
for col in range(1, areas.max_column+1):
    areas[get_column_letter(col)+str(areas.max_row)].font = Font(bold=True)
    areas[get_column_letter(col)+str(areas.max_row)].fill = PatternFill(
        fill_type='solid', start_color=gray, end_color=gray)

# AREAS SUMMARY BY MATERIAL
for each in job['materials']:
    # CREATE HEADERS
    abmmr = abm.max_row
    if abmmr == 1:
        abmmr = -1
    abm['A'+str(abmmr+2)] = each['material'] + ' Summary'
    abm[mat_area+str(abmmr+3)] = 'Area'
    abm[mat_priority+str(abmmr+3)] = 'Priority #'
    abm[mat_spools+str(abmmr+3)] = 'Total Spools'
    abm[mat_on_hold+str(abmmr+3)] = 'On Hold'
    abm[mat_workable+str(abmmr+3)] = 'Workable'
    abm[mat_not_workable+str(abmmr+3)] = 'Not Workable'
    abm[mat_weldout+str(abmmr+3)] = 'Welded Out'
    abm[mat_weldout_remaining+str(abmmr+3)] = 'Remaining to Weld Out'
    abm[mat_stc+str(abmmr+3)] = 'Shipped To Paint'
    abm[mat_delivered+str(abmmr+3)] = 'Delivered'
    abm[mat_delivered_remaining+str(abmmr+3)] = 'Remaining to Deliver'
    abm[mat_perc_workable+str(abmmr+3)] = 'Workable %'
    abm[mat_perc_weldout+str(abmmr+3)] = 'Weld Out %'
    abm[mat_perc_delivered+str(abmmr+3)] = 'Delivered %'
    for area in each['areas']:
        abm[mat_area+str(abm.max_row+1)] = area['area']
        abm[mat_priority+str(abm.max_row)] = area['priority']
        abm[mat_spools+str(abm.max_row)] = area['spools']
        abm[mat_on_hold+str(abm.max_row)] = area['on_hold']
        abm[mat_workable+str(abm.max_row)] = area['workable']
        abm[mat_not_workable+str(abm.max_row)
            ] = area['spools']-area['workable']
        abm[mat_weldout+str(abm.max_row)] = area['weldout']
        abm[mat_weldout_remaining +
            str(abm.max_row)] = area['spools']-area['weldout']
        abm[mat_stc +
            str(abm.max_row)] = area['stc']
        abm[mat_delivered+str(abm.max_row)
            ] = area['delivered']
        abm[mat_delivered_remaining +
            str(abm.max_row)] = area['spools']-area['delivered']
        abm[mat_perc_workable +
            str(abm.max_row)] = str(area['workable_perc']) + '%'
        abm[mat_perc_weldout +
            str(abm.max_row)] = str(area['weldout_perc']) + '%'
        abm[mat_perc_delivered +
            str(abm.max_row)] = str(area['delivered_perc']) + '%'

    # TOTALS
    abm['A'+str(abm.max_row+1)] = 'TOTALS'
    for col in range(1, abm.max_column+1):
        abm[get_column_letter(col)+str(abm.max_row)].font = Font(bold=True)
        abm[get_column_letter(col)+str(abm.max_row)].fill = PatternFill(
            fill_type='solid', start_color=gray, end_color=gray)
        abm[mat_spools+str(abm.max_row)] = each['total']
        abm[mat_on_hold+str(abm.max_row)] = each['on_hold']
        abm[mat_workable+str(abm.max_row)] = each['workable']
        abm[mat_not_workable+str(abm.max_row)
            ] = each['total']-each['workable']
        abm[mat_weldout+str(abm.max_row)] = each['weldout']
        abm[mat_weldout_remaining +
            str(abm.max_row)] = each['total']-each['weldout']
        abm[mat_stc+str(abm.max_row)] = each['stc']
        abm[mat_delivered+str(abm.max_row)] = each['delivered']
        abm[mat_delivered_remaining +
            str(abm.max_row)] = each['total']-each['delivered']
        abm[mat_perc_workable+str(abm.max_row)
            ] = str(round((each['workable']/each['total'])*100))+"%"
        abm[mat_perc_weldout+str(abm.max_row)
            ] = str(round((each['weldout']/each['total'])*100))+"%"
        abm[mat_perc_delivered+str(abm.max_row)
            ] = str(round((each['delivered']/each['total'])*100))+"%"
        abm.row_dimensions[abm.max_row].height = 20

# # SPOOL LOCAITON
# print('')
# print('Adding spool locations..')

# loc['B1'] = job['location_nw']
# loc['B2'] = job['location_w']
# loc['B3'] = job['location_i']
# loc['B4'] = job['location_p']
# loc['B5'] = job['location_wo']
# loc['B6'] = job['location_stp']
# loc['B7'] = job['location_d']

# SHORTS BY SCOPE
print('')
print('Adding shorts by scope..')

shbs['B1'] = str(job['total']) + " Spools"
shbs['C1'] = str(job['workable']) + " Workable"
shbs['D1'] = str(job['issued']) + " Issued"
shbs['B4'] = job['count_shorts']['total']['valves']['performance']
shbs['B5'] = job['count_shorts']['total']['flanges']['performance']
shbs['B6'] = job['count_shorts']['total']['fittings']['performance']
shbs['B7'] = job['count_shorts']['total']['supports']['performance']
shbs['B10'] = job['count_shorts']['total']['pipe']['performance']
shbs['C4'] = job['count_shorts']['total']['valves']['client']
shbs['C5'] = job['count_shorts']['total']['flanges']['client']
shbs['C6'] = job['count_shorts']['total']['fittings']['client']
shbs['C7'] = job['count_shorts']['total']['supports']['client']
shbs['C10'] = job['count_shorts']['total']['pipe']['client']
shbs['D4'] = job['count_shorts']['total']['valves']['other']
shbs['D5'] = job['count_shorts']['total']['flanges']['other']
shbs['D6'] = job['count_shorts']['total']['fittings']['other']
shbs['D7'] = job['count_shorts']['total']['supports']['other']
shbs['D10'] = job['count_shorts']['total']['pipe']['other']

# PURCHASED BY SCOPE
print('')
print('Adding purchased by scope..')

pbs['B1'] = str(job['total']) + " Spools"
pbs['C1'] = str(job['workable']) + " Workable"
pbs['D1'] = str(job['issued']) + " Issued"
pbs['B4'] = job['count_shorts']['purchased']['valves']['performance']
pbs['B5'] = job['count_shorts']['purchased']['flanges']['performance']
pbs['B6'] = job['count_shorts']['purchased']['fittings']['performance']
pbs['B7'] = job['count_shorts']['purchased']['supports']['performance']
pbs['B10'] = job['count_shorts']['purchased']['pipe']['performance']
pbs['C4'] = job['count_shorts']['purchased']['valves']['client']
pbs['C5'] = job['count_shorts']['purchased']['flanges']['client']
pbs['C6'] = job['count_shorts']['purchased']['fittings']['client']
pbs['C7'] = job['count_shorts']['purchased']['supports']['client']
pbs['C10'] = job['count_shorts']['purchased']['pipe']['client']
pbs['D4'] = job['count_shorts']['purchased']['valves']['other']
pbs['D5'] = job['count_shorts']['purchased']['flanges']['other']
pbs['D6'] = job['count_shorts']['purchased']['fittings']['other']
pbs['D7'] = job['count_shorts']['purchased']['supports']['other']
pbs['D10'] = job['count_shorts']['purchased']['pipe']['other']

# NO MATERIAL BY SCOPE
print('')
print('Adding no material by scope..')

nmbs['B1'] = str(job['total']) + " Spools"
nmbs['C1'] = str(job['workable']) + " Workable"
nmbs['D1'] = str(job['issued']) + " Issued"
nmbs['B4'] = job['count_shorts']['no_material']['valves']['performance']
nmbs['B5'] = job['count_shorts']['no_material']['flanges']['performance']
nmbs['B6'] = job['count_shorts']['no_material']['fittings']['performance']
nmbs['B7'] = job['count_shorts']['no_material']['supports']['performance']
nmbs['B10'] = job['count_shorts']['no_material']['pipe']['performance']
nmbs['C4'] = job['count_shorts']['no_material']['valves']['client']
nmbs['C5'] = job['count_shorts']['no_material']['flanges']['client']
nmbs['C6'] = job['count_shorts']['no_material']['fittings']['client']
nmbs['C7'] = job['count_shorts']['no_material']['supports']['client']
nmbs['C10'] = job['count_shorts']['no_material']['pipe']['client']
nmbs['D4'] = job['count_shorts']['no_material']['valves']['other']
nmbs['D5'] = job['count_shorts']['no_material']['flanges']['other']
nmbs['D6'] = job['count_shorts']['no_material']['fittings']['other']
nmbs['D7'] = job['count_shorts']['no_material']['supports']['other']
nmbs['D10'] = job['count_shorts']['no_material']['pipe']['other']

# NO MATERIAL BY SCOPE
print('')
print('Adding no spools missing valve only..')

mvo['B3'] = job['count_shorts']['missing_valve_only']['performance']
mvo['C3'] = job['count_shorts']['missing_valve_only']['client']
mvo['D3'] = job['count_shorts']['missing_valve_only']['other']

# SPOOLS BY SCOPE
print('')
print('Adding spools by scope..')

spbs['B1'] = str(job['total']) + " Spools"
spbs['C1'] = str(job['workable']) + " Workable"
spbs['D1'] = str(job['issued']) + " Issued"
spbs['B4'] = job['spools_by_scope']['valves']['performance']
spbs['B5'] = job['spools_by_scope']['pipe']['performance']
spbs['B6'] = job['spools_by_scope']['flanges']['performance']
spbs['B7'] = job['spools_by_scope']['fittings']['performance']
spbs['B8'] = job['spools_by_scope']['supports']['performance']
spbs['C4'] = job['spools_by_scope']['valves']['client']
spbs['C5'] = job['spools_by_scope']['pipe']['client']
spbs['C6'] = job['spools_by_scope']['flanges']['client']
spbs['C7'] = job['spools_by_scope']['fittings']['client']
spbs['C8'] = job['spools_by_scope']['supports']['client']
spbs['D4'] = job['spools_by_scope']['valves']['other']
spbs['D5'] = job['spools_by_scope']['pipe']['other']
spbs['D6'] = job['spools_by_scope']['flanges']['other']
spbs['D7'] = job['spools_by_scope']['fittings']['other']
spbs['D8'] = job['spools_by_scope']['supports']['other']
spbs['E9'] = job['issued']
spbs['E10'] = job['workable_not_issued']
spbs['E11'] = job['issued_missing_item']
spbs['E12'] = job['on_hold_no_shorts']
spbs['E13'] = job['spools_by_scope']['discrepancies']
spbs['E14'] = job['total']

# DISCREPANCIES
print('')
print('Adding discrepancies lists..')

for i in range(len(job['discrepancies']['fc_iss'])):
    disc['A'+str(i+2)] = job['discrepancies']['fc_iss'][i]['spool']
for i in range(len(job['discrepancies']['notfc_notiss'])):
    disc['B'+str(i+2)] = job['discrepancies']['notfc_not_iss'][i]['spool']
for i in range(len(job['discrepancies']['fc_not_ll'])):
    disc['C'+str(i+2)] = job['discrepancies']['fc_not_ll'][i]['spool']
for i in range(len(job['discrepancies']['sr_not_ll'])):
    disc['D'+str(i+2)] = job['discrepancies']['sr_not_ll'][i]['piecemark']

for row in range(2, disc.max_row+1):
    for col in range(1, disc.max_column+1):
        disc[get_column_letter(
            col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')

# STYLES #####################
# AREAS

# PRINTOUT
for row in range(29, printout.max_row+1):
    for col in range(1, printout.max_column+1):
        printout[get_column_letter(
            col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')

# AREA TABS
areasmc = areas.max_column
areasmr = areas.max_row

for row in range(1, areasmr+1):
    areas.row_dimensions[row].height = 15
    for col in range(1, areasmc+1):
        areas.column_dimensions[str(get_column_letter(col))].width = 13
        areas[str(get_column_letter(col))+str(row)
              ].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

areas.row_dimensions[1].height = 37
areas.row_dimensions[2].height = 37

areas.merge_cells('A1:'+str(get_column_letter(areas.max_column))+'1')

areas['A1'].font = Font(bold=True, sz=20, color=white)
areas['A1'].fill = PatternFill(
    fill_type='solid', start_color=black, end_color=black)

for col in range(1, areas.max_column+1):
    areas[get_column_letter(col)+'2'].font = Font(bold=True, color=white)
    areas[get_column_letter(col)+'2'].fill = PatternFill(
        fill_type='solid', start_color=header_blue, end_color=header_blue)

# ABM SHEET
# SIZE AND ALIGN ALL CELLS
if len(job['areas']) > 0:
    for row in range(1, abm.max_row+1):
        abm.row_dimensions[row].height = 15
        for col in range(1, abm.max_column+1):
            abm.column_dimensions[str(get_column_letter(col))].width = 13
            abm[str(get_column_letter(col))+str(row)
                ].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

        if abm['A'+str(row)].value == 'Area':
            abm.merge_cells('A'+str(row-1)+':'+mat_perc_delivered+str(row-1))
            abm.row_dimensions[row].height = 37
            abm.row_dimensions[row-1].height = 37
            abm['A'+str(row-1)].font = Font(bold=True, sz=20, color=white)
            abm['A'+str(row-1)].fill = PatternFill(
                fill_type='solid', start_color=black, end_color=black)
            carbon_headers = ['A'+str(row), 'B'+str(row), 'C'+str(row), 'D'+str(row), 'E'+str(row), 'F'+str(row), 'G'+str(row), 'H'+str(row),
                              'I'+str(row), 'J'+str(row), 'K'+str(row), 'L'+str(row), 'M'+str(row), 'N'+str(row)]
            for cell in carbon_headers:
                abm[cell].font = Font(bold=True, color=white)
                abm[cell].fill = PatternFill(
                    fill_type='solid', start_color=header_blue, end_color=header_blue)


# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

# SAVE AS NEW WORKBOOK
wb.save(filename='database/'+str(job_number)+'/JobSummary.xlsx')

print('Summary Complete!')
