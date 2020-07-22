import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, colors, Alignment, PatternFill, Border
from datetime import date, datetime
import json

client_name = sys.argv[2]
job_number = sys.argv[1]

# LOAD WORKBOOKS
wb = load_workbook(filename='database/Templates/JobSumTemplate.xlsx')

# # PULL IN JOB JSON FILE
# with open('jobs/'+str(job_number)+'/job.json', 'r') as f:
#     job_json = json.load(f)

# job = job_json

# # SET ACTIVATIONS
# printout = wb['PRINTOUT']
# areas = wb['Areas Summary']
# abm = wb['Areas Summary (By Material)']
# loc = wb['Spool Location']
# shbs = wb['Shorts by Scope']
# pbs = wb['Purchased by Scope']
# nmbs = wb['No Material by Scope']
# mvo = wb['Spool Missing Valve Only']
# los = wb['List of Shorts']
# spbs = wb['Spools by Scope']
# spools = wb['List of Spools']
# disc = wb['Discrepancies']

# print('')
# print('Adding spools to summary..')

# # COLORS
# black = '404040'
# gray = 'd9d9d9'
# header_blue = '538dd5'
# data_blue = 'e7eff9'
# header_red = 'd27d7c'
# data_red = 'fbf3f3'
# header_green = 'abc56d'
# data_green = 'f4f8ec'
# header_orange = 'f79443'
# data_orange = 'fdf0e7'
# white = 'ffffff'
# discrepancies = 'ffb3bc'

# # COLUMNS
# mat_area = 'A'
# mat_priority = 'B'
# mat_spools = 'C'
# mat_on_hold = 'D'
# mat_workable = 'E'
# mat_not_workable = 'F'
# mat_weld_out = 'G'
# mat_weld_out_remaining = 'H'
# mat_stc = 'I'
# mat_delivered = 'J'
# mat_delivered_remaining = 'K'
# mat_perc_workable = 'L'
# mat_perc_weld_out = 'M'
# mat_perc_delivered = 'N'

# # AREAS SUMMARY PRINTOUT

# # COLUMNS
# mat_area = 'A'
# mat_priority = 'B'
# mat_spools = 'C'
# mat_on_hold = 'D'
# mat_workable = 'E'
# mat_not_workable = 'F'
# mat_weld_out = 'G'
# mat_weld_out_remaining = 'H'
# mat_stc = 'I'
# mat_delivered = 'J'
# mat_delivered_remaining = 'K'
# mat_perc_workable = 'L'
# mat_perc_weld_out = 'M'
# mat_perc_delivered = 'N'

# printout['A1'] = client_name + ': ' + job_number
# printout['C1'] = str(job['total_spools']) + " Spools"
# printout['F1'] = str(job['total_workable']) + " Workable"
# printout['I1'] = str(job['total_issued']) + " Issued"

# printout['A28'] = job_number + ' Areas Summary'
# for area in job['areas']:
#     printout[mat_area+str(printout.max_row+1)] = area['area']
#     printout[mat_priority+str(printout.max_row)] = area['priority']
#     printout[mat_spools+str(printout.max_row)] = area['total_spools']
#     printout[mat_on_hold+str(printout.max_row)] = area['total_on_hold']
#     printout[mat_workable+str(printout.max_row)] = area['total_workable']
#     printout[mat_not_workable+str(printout.max_row)
#              ] = area['total_spools']-area['total_workable']
#     printout[mat_weld_out+str(printout.max_row)] = area['total_weld_out']
#     printout[mat_weld_out_remaining +
#              str(printout.max_row)] = area['total_spools']-area['total_weld_out']
#     printout[mat_stc +
#              str(printout.max_row)] = area['total_stc']
#     printout[mat_delivered+str(printout.max_row)] = area['total_delivered']
#     printout[mat_delivered_remaining +
#              str(printout.max_row)] = area['total_spools']-area['total_delivered']
#     printout[mat_perc_workable+str(printout.max_row)
#              ] = str(round((area['total_workable']/area['total_spools'])*100))+"%"
#     printout[mat_perc_weld_out+str(printout.max_row)
#              ] = str(round((area['total_weld_out']/area['total_spools'])*100))+"%"
#     printout[mat_perc_delivered+str(printout.max_row)
#              ] = str(round((area['total_delivered']/area['total_spools'])*100))+"%"

# # TOTALS
# printout['A'+str(printout.max_row+1)] = 'TOTALS'
# printout[mat_spools+str(printout.max_row)] = job['total_spools']
# printout[mat_on_hold+str(printout.max_row)] = job['total_on_hold']
# printout[mat_workable+str(printout.max_row)] = job['total_workable']
# printout[mat_not_workable+str(printout.max_row)
#          ] = job['total_spools']-job['total_workable']
# printout[mat_weld_out+str(printout.max_row)] = job['total_weld_out']
# printout[mat_weld_out_remaining +
#          str(printout.max_row)] = job['total_spools']-job['total_weld_out']
# printout[mat_stc+str(printout.max_row)] = job['total_stc']
# printout[mat_delivered+str(printout.max_row)] = job['total_delivered']
# printout[mat_delivered_remaining +
#          str(printout.max_row)] = job['total_spools']-job['total_delivered']
# printout[mat_perc_workable+str(printout.max_row)
#          ] = str(round((job['total_workable']/job['total_spools'])*100))+"%"
# printout[mat_perc_weld_out+str(printout.max_row)
#          ] = str(round((job['total_weld_out']/job['total_spools'])*100))+"%"
# printout[mat_perc_delivered+str(printout.max_row)
#          ] = str(round((job['total_delivered']/job['total_spools'])*100))+"%"
# printout.row_dimensions[printout.max_row].height = 20

# # FILL TOTALS ROW
# for col in range(1, printout.max_column+1):
#     printout[get_column_letter(col)+str(printout.max_row)
#              ].font = Font(bold=True)
#     printout[get_column_letter(col)+str(printout.max_row)].fill = PatternFill(
#         fill_type='solid', start_color=gray, end_color=gray)

# # AREAS SUMMARY TAB
# areas['A1'] = 'Areas Summary'
# areas[mat_area+'2'] = 'Area'
# areas[mat_priority+'2'] = 'Priority'
# areas[mat_spools+'2'] = 'Total Spools'
# areas[mat_on_hold+'2'] = 'On Hold'
# areas[mat_workable+'2'] = 'Workable'
# areas[mat_not_workable+'2'] = 'Not Workable'
# areas[mat_weld_out+'2'] = 'Welded Out'
# areas[mat_weld_out_remaining+'2'] = 'Remaining to Weld Out'
# areas[mat_stc+'2'] = 'Shipped To Paint'
# areas[mat_delivered+'2'] = 'Delivered'
# areas[mat_delivered_remaining+'2'] = 'Remaining to Deliver'
# areas[mat_perc_workable+'2'] = 'Workable %'
# areas[mat_perc_weld_out+'2'] = 'Weld Out %'
# areas[mat_perc_delivered+'2'] = 'Delivered %'

# for area in job['areas']:
#     areas[mat_area+str(areas.max_row+1)] = area['area']
#     areas[mat_priority+str(areas.max_row)] = area['priority']
#     areas[mat_spools+str(areas.max_row)] = area['total_spools']
#     areas[mat_on_hold+str(areas.max_row)] = area['total_on_hold']
#     areas[mat_workable+str(areas.max_row)] = area['total_workable']
#     areas[mat_not_workable+str(areas.max_row)
#           ] = area['total_spools']-area['total_workable']
#     areas[mat_weld_out+str(areas.max_row)] = area['total_weld_out']
#     areas[mat_weld_out_remaining +
#           str(areas.max_row)] = area['total_spools']-area['total_weld_out']
#     areas[mat_stc +
#           str(areas.max_row)] = area['total_stc']
#     areas[mat_delivered+str(areas.max_row)] = area['total_delivered']
#     areas[mat_delivered_remaining +
#           str(areas.max_row)] = area['total_spools']-area['total_delivered']
#     areas[mat_perc_workable+str(areas.max_row)
#           ] = str(round((area['total_workable']/area['total_spools'])*100))+"%"
#     areas[mat_perc_weld_out+str(areas.max_row)
#           ] = str(round((area['total_weld_out']/area['total_spools'])*100))+"%"
#     areas[mat_perc_delivered+str(areas.max_row)
#           ] = str(round((area['total_delivered']/area['total_spools'])*100))+"%"

# # TOTALS
# areas['A'+str(areas.max_row+1)] = 'TOTALS'
# areas[mat_spools+str(areas.max_row)] = job['total_spools']
# areas[mat_on_hold+str(areas.max_row)] = job['total_on_hold']
# areas[mat_workable+str(areas.max_row)] = job['total_workable']
# areas[mat_not_workable+str(areas.max_row)
#       ] = job['total_spools']-job['total_workable']
# areas[mat_weld_out+str(areas.max_row)] = job['total_weld_out']
# areas[mat_weld_out_remaining +
#       str(areas.max_row)] = job['total_spools']-job['total_weld_out']
# areas[mat_stc+str(areas.max_row)] = job['total_stc']
# areas[mat_delivered+str(areas.max_row)] = job['total_delivered']
# areas[mat_delivered_remaining +
#       str(areas.max_row)] = job['total_spools']-job['total_delivered']
# areas[mat_perc_workable+str(areas.max_row)
#       ] = str(round((job['total_workable']/job['total_spools'])*100))+"%"
# areas[mat_perc_weld_out+str(areas.max_row)
#       ] = str(round((job['total_weld_out']/job['total_spools'])*100))+"%"
# areas[mat_perc_delivered+str(areas.max_row)
#       ] = str(round((job['total_delivered']/job['total_spools'])*100))+"%"
# areas.row_dimensions[areas.max_row].height = 20

# # FILL TOTALS ROW
# for col in range(1, areas.max_column+1):
#     areas[get_column_letter(col)+str(areas.max_row)].font = Font(bold=True)
#     areas[get_column_letter(col)+str(areas.max_row)].fill = PatternFill(
#         fill_type='solid', start_color=gray, end_color=gray)

# # AREAS SUMMARY BY MATERIAL
# for each in job['materials_list']:
#     # CREATE HEADERS
#     abmmr = abm.max_row
#     if abmmr == 1:
#         abmmr = -1
#     abm['A'+str(abmmr+2)] = each + ' Summary'
#     abm[mat_area+str(abmmr+3)] = 'Area'
#     abm[mat_priority+str(abmmr+3)] = 'Priority #'
#     abm[mat_spools+str(abmmr+3)] = 'Total Spools'
#     abm[mat_on_hold+str(abmmr+3)] = 'On Hold'
#     abm[mat_workable+str(abmmr+3)] = 'Workable'
#     abm[mat_not_workable+str(abmmr+3)] = 'Not Workable'
#     abm[mat_weld_out+str(abmmr+3)] = 'Welded Out'
#     abm[mat_weld_out_remaining+str(abmmr+3)] = 'Remaining to Weld Out'
#     abm[mat_stc+str(abmmr+3)] = 'Shipped To Paint'
#     abm[mat_delivered+str(abmmr+3)] = 'Delivered'
#     abm[mat_delivered_remaining+str(abmmr+3)] = 'Remaining to Deliver'
#     abm[mat_perc_workable+str(abmmr+3)] = 'Workable %'
#     abm[mat_perc_weld_out+str(abmmr+3)] = 'Weld Out %'
#     abm[mat_perc_delivered+str(abmmr+3)] = 'Delivered %'

#     # MATERIAL AREA INFO
#     for area in job['areas']:
#         for material in area['materials']:
#             if material['name'] == each:
#                 abm[mat_area+str(abm.max_row+1)] = area['area']
#                 abm[mat_priority+str(abm.max_row)] = area['priority']
#                 abm[mat_spools+str(abm.max_row)] = material['total_spools']
#                 abm[mat_on_hold+str(abm.max_row)] = material['total_on_hold']
#                 abm[mat_workable+str(abm.max_row)] = material['total_workable']
#                 abm[mat_not_workable+str(abm.max_row)
#                     ] = material['total_spools']-material['total_workable']
#                 abm[mat_weld_out+str(abm.max_row)] = material['total_weld_out']
#                 abm[mat_weld_out_remaining +
#                     str(abm.max_row)] = material['total_spools']-material['total_weld_out']
#                 abm[mat_stc +
#                     str(abm.max_row)] = material['total_stc']
#                 abm[mat_delivered+str(abm.max_row)
#                     ] = material['total_delivered']
#                 abm[mat_delivered_remaining +
#                     str(abm.max_row)] = material['total_spools']-material['total_delivered']
#                 abm[mat_perc_workable +
#                     str(abm.max_row)] = str(round((material['total_workable']/material['total_spools'])*100))+"%"
#                 abm[mat_perc_weld_out +
#                     str(abm.max_row)] = str(round((material['total_weld_out']/material['total_spools'])*100))+"%"
#                 abm[mat_perc_delivered +
#                     str(abm.max_row)] = str(round((material['total_delivered']/material['total_spools'])*100))+"%"

#     # TOTALS
#     abm['A'+str(abm.max_row+1)] = 'TOTALS'
#     for col in range(1, abm.max_column+1):
#         abm[get_column_letter(col)+str(abm.max_row)].font = Font(bold=True)
#         abm[get_column_letter(col)+str(abm.max_row)].fill = PatternFill(
#             fill_type='solid', start_color=gray, end_color=gray)
#         for one in job['materials']:
#             if one['name'] == each:
#                 abm[mat_spools+str(abm.max_row)] = one['total_spools']
#                 abm[mat_on_hold+str(abm.max_row)] = one['total_on_hold']
#                 abm[mat_workable+str(abm.max_row)] = one['workable']
#                 abm[mat_not_workable+str(abm.max_row)
#                     ] = one['total_spools']-one['workable']
#                 abm[mat_weld_out+str(abm.max_row)] = one['weld_out']
#                 abm[mat_weld_out_remaining +
#                     str(abm.max_row)] = one['total_spools']-one['weld_out']
#                 abm[mat_stc+str(abm.max_row)] = one['stc']
#                 abm[mat_delivered+str(abm.max_row)] = one['delivered']
#                 abm[mat_delivered_remaining +
#                     str(abm.max_row)] = one['total_spools']-one['delivered']
#                 abm[mat_perc_workable+str(abm.max_row)
#                     ] = str(round((one['workable']/one['total_spools'])*100))+"%"
#                 abm[mat_perc_weld_out+str(abm.max_row)
#                     ] = str(round((one['weld_out']/one['total_spools'])*100))+"%"
#                 abm[mat_perc_delivered+str(abm.max_row)
#                     ] = str(round((one['delivered']/one['total_spools'])*100))+"%"
#                 abm.row_dimensions[abm.max_row].height = 20

# # SPOOL LOCAITON
# print('')
# print('Adding spool locations..')

# loc['B1'] = job['location_nw']
# loc['B2'] = job['location_w']
# loc['B3'] = job['location_i']
# loc['B4'] = job['location_p']
# loc['B5'] = job['location_wo']
# loc['B6'] = job['location_stp']
# loc['B7'] = job['location_d']

# # SHORTS BY SCOPE
# print('')
# print('Adding shorts by scope..')

# shbs['B1'] = str(job['total_spools']) + " Spools"
# shbs['C1'] = str(job['total_workable']) + " Workable"
# shbs['D1'] = str(job['total_issued']) + " Issued"
# shbs['B4'] = job['p_items_valves']
# shbs['B5'] = job['p_items_flanges']
# shbs['B6'] = job['p_items_fittings']
# shbs['B7'] = job['p_items_supports']
# shbs['B10'] = job['p_items_pipe']
# shbs['C4'] = job['c_items_valves']
# shbs['C5'] = job['c_items_flanges']
# shbs['C6'] = job['c_items_fittings']
# shbs['C7'] = job['c_items_supports']
# shbs['C10'] = job['c_items_pipe']
# shbs['D4'] = job['o_items_valves']
# shbs['D5'] = job['o_items_flanges']
# shbs['D6'] = job['o_items_fittings']
# shbs['D7'] = job['o_items_supports']
# shbs['D10'] = job['o_items_pipe']

# total_shorts = 0
# for short in job['shorts_not_ll']:
#     total_shorts += short['quantity']
# shbs['E12'] = total_shorts

# # PURCHASED BY SCOPE
# print('')
# print('Adding purchased by scope..')

# pbs['B1'] = str(job['total_spools']) + " Spools"
# pbs['C1'] = str(job['total_workable']) + " Workable"
# pbs['D1'] = str(job['total_issued']) + " Issued"
# pbs['B4'] = job['p_purchased_valves']
# pbs['B5'] = job['p_purchased_flanges']
# pbs['B6'] = job['p_purchased_fittings']
# pbs['B7'] = job['p_purchased_supports']
# pbs['B10'] = job['p_purchased_pipe']
# pbs['C4'] = job['c_purchased_valves']
# pbs['C5'] = job['c_purchased_flanges']
# pbs['C6'] = job['c_purchased_fittings']
# pbs['C7'] = job['c_purchased_supports']
# pbs['C10'] = job['c_purchased_pipe']
# pbs['D4'] = job['o_purchased_valves']
# pbs['D5'] = job['o_purchased_flanges']
# pbs['D6'] = job['o_purchased_fittings']
# pbs['D7'] = job['o_purchased_supports']
# pbs['D10'] = job['o_purchased_pipe']

# # NO MATERIAL BY SCOPE
# print('')
# print('Adding no material by scope..')

# nmbs['B1'] = str(job['total_spools']) + " Spools"
# nmbs['C1'] = str(job['total_workable']) + " Workable"
# nmbs['D1'] = str(job['total_issued']) + " Issued"
# nmbs['B4'] = job['p_nomaterial_valves']
# nmbs['B5'] = job['p_nomaterial_flanges']
# nmbs['B6'] = job['p_nomaterial_fittings']
# nmbs['B7'] = job['p_nomaterial_supports']
# nmbs['B10'] = job['p_nomaterial_pipe']
# nmbs['C4'] = job['c_nomaterial_valves']
# nmbs['C5'] = job['c_nomaterial_flanges']
# nmbs['C6'] = job['c_nomaterial_fittings']
# nmbs['C7'] = job['c_nomaterial_supports']
# nmbs['C10'] = job['c_nomaterial_pipe']
# nmbs['D4'] = job['o_nomaterial_valves']
# nmbs['D5'] = job['o_nomaterial_flanges']
# nmbs['D6'] = job['o_nomaterial_fittings']
# nmbs['D7'] = job['o_nomaterial_supports']
# nmbs['D10'] = job['o_nomaterial_pipe']

# # NO MATERIAL BY SCOPE
# print('')
# print('Adding no spools missing valve only..')

# mvo['B3'] = job['valve_only_p']
# mvo['C3'] = job['valve_only_c']
# mvo['D3'] = job['valve_only_o']

# # LIST OF SHORTS
# print('')
# print('Adding list of shorts..')

# for spool in job['spools']:
#     for comp in spool['comps']:
#         if comp['status'] != 'Complete':
#             los['A'+str(los.max_row+1)] = spool['spool']
#             los['B'+str(los.max_row)] = comp['scope']
#             los['C'+str(los.max_row)] = comp['tag']
#             los['D'+str(los.max_row)] = comp['status']
#             los['E'+str(los.max_row)] = comp['item']
#             los['F'+str(los.max_row)] = comp['quantity']
#             los['G'+str(los.max_row)] = spool['material']
#             los['H'+str(los.max_row)] = spool['priority']

# for short in job['shorts_not_ll']:
#     los['A'+str(los.max_row+1)] = short['spool']
#     los['B'+str(los.max_row)] = short['scope']
#     los['C'+str(los.max_row)] = short['tag']
#     los['D'+str(los.max_row)] = short['status']
#     los['E'+str(los.max_row)] = short['item']
#     los['F'+str(los.max_row)] = short['quantity']
#     los['G'+str(los.max_row)] = None
#     los['H'+str(los.max_row)] = None

#     shorts_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
#     for each in shorts_columns:
#         los[each+str(los.max_row)].fill = PatternFill(
#             fill_type='solid', start_color=discrepancies, end_color=discrepancies)

# for row in range(2, los.max_row+1):
#     for col in range(1, los.max_column+1):
#         los[get_column_letter(
#             col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')

# # SPOOLS BY SCOPE
# print('')
# print('Adding spools by scope..')

# spbs['B1'] = str(job['total_spools']) + " Spools"
# spbs['C1'] = str(job['total_workable']) + " Workable"
# spbs['D1'] = str(job['total_issued']) + " Issued"
# spbs['B4'] = job['hold_p_valves']
# spbs['B5'] = job['hold_p_flanges']
# spbs['B6'] = job['hold_p_fittings']
# spbs['B7'] = job['hold_p_pipe']
# spbs['B8'] = job['hold_p_supports']
# spbs['C4'] = job['hold_c_valves']
# spbs['C5'] = job['hold_c_flanges']
# spbs['C6'] = job['hold_c_fittings']
# spbs['C7'] = job['hold_c_pipe']
# spbs['C8'] = job['hold_c_supports']
# spbs['D4'] = job['hold_o_valves']
# spbs['D5'] = job['hold_o_flanges']
# spbs['D6'] = job['hold_o_fittings']
# spbs['D7'] = job['hold_o_pipe']
# spbs['D8'] = job['hold_o_supports']
# spbs['E9'] = job['total_issued']
# spbs['E10'] = job['total_workable']-job['total_issued']
# spbs['E12'] = job['total_spools']


# # ALL SPOOLS
# for spool in job['spools']:
#     spools['A'+str(spools.max_row+1)] = spool['spool']
#     spools['B'+str(spools.max_row)] = spool['multiplier']
#     spools['C'+str(spools.max_row)] = spool['area']
#     spools['D'+str(spools.max_row)] = spool['priority']
#     spools['E'+str(spools.max_row)] = spool['material']
#     spools['F'+str(spools.max_row)] = spool['shop']
#     spools['G'+str(spools.max_row)] = spool['lifespan']
#     spools['H'+str(spools.max_row)] = spool['location']
#     spools['I'+str(spools.max_row)] = spool['dormant']
#     spools['J'+str(spools.max_row)] = spool['on_hold']
#     spools['K'+str(spools.max_row)] = spool['workable']
#     spools['L'+str(spools.max_row)] = spool['issued']
#     spools['M'+str(spools.max_row)] = spool['date_pull']
#     spools['N'+str(spools.max_row)] = spool['weld_out']
#     spools['O'+str(spools.max_row)] = spool['stc']
#     spools['P'+str(spools.max_row)] = spool['delivered']
#     spools['Q'+str(spools.max_row)] = spool['total_pipe']
#     spools['R'+str(spools.max_row)] = spool['total_items']
#     spools['S'+str(spools.max_row)] = spool['missing_pipe']
#     spools['T'+str(spools.max_row)] = spool['missing_items']
#     spools['U'+str(spools.max_row)] = spool['missing_valves']
#     spools['V'+str(spools.max_row)] = spool['missing_flanges']
#     spools['W'+str(spools.max_row)] = spool['missing_fittings']
#     spools['X'+str(spools.max_row)] = spool['missing_supports']
#     spools['Y'+str(spools.max_row)] = spool['hold']
#     spools['Z'+str(spools.max_row)] = spool['scope']

# for row in range(2, spools.max_row+1):
#     for col in range(1, spools.max_column+1):
#         spools[get_column_letter(
#             col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')


# # DISCREPANCIES
# print('')
# print('Adding discrepancies lists..')

# for i in range(len(job['fc_issued'])):
#     disc['A'+str(i+2)] = job['fc_issued'][i]
# for i in range(len(job['not_fc_not_issued'])):
#     disc['B'+str(i+2)] = job['not_fc_not_issued'][i]
# for i in range(len(job['fc_not_ll'])):
#     disc['C'+str(i+2)] = job['fc_not_ll'][i]
# for i in range(len(job['sr_not_ll'])):
#     disc['D'+str(i+2)] = job['sr_not_ll'][i]
# for i in range(len(job['on_hold_dis'])):
#     disc['D'+str(i+2)] = job['on_hold_dis'][i]

# for row in range(2, disc.max_row+1):
#     for col in range(1, disc.max_column+1):
#         disc[get_column_letter(
#             col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')

# # AREAS

# # PRINTOUT
# for row in range(29, printout.max_row+1):
#     for col in range(1, printout.max_column+1):
#         printout[get_column_letter(
#             col)+str(row)].alignment = Alignment(vertical='center', horizontal='center')

# # AREA TABS
# areasmc = areas.max_column
# areasmr = areas.max_row

# for row in range(1, areasmr+1):
#     areas.row_dimensions[row].height = 15
#     for col in range(1, areasmc+1):
#         areas.column_dimensions[str(get_column_letter(col))].width = 13
#         areas[str(get_column_letter(col))+str(row)
#               ].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

# areas.row_dimensions[1].height = 37
# areas.row_dimensions[2].height = 37

# areas.merge_cells('A1:'+str(get_column_letter(areas.max_column))+'1')

# areas['A1'].font = Font(bold=True, sz=20, color=white)
# areas['A1'].fill = PatternFill(
#     fill_type='solid', start_color=black, end_color=black)

# for col in range(1, areas.max_column+1):
#     areas[get_column_letter(col)+'2'].font = Font(bold=True, color=white)
#     areas[get_column_letter(col)+'2'].fill = PatternFill(
#         fill_type='solid', start_color=header_blue, end_color=header_blue)

# # ABM SHEET
# # SIZE AND ALIGN ALL CELLS
# if len(job['areas']) > 0:
#     for row in range(1, abm.max_row+1):
#         abm.row_dimensions[row].height = 15
#         for col in range(1, abm.max_column+1):
#             abm.column_dimensions[str(get_column_letter(col))].width = 13
#             abm[str(get_column_letter(col))+str(row)
#                 ].alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

#         if abm['A'+str(row)].value == 'Area':
#             abm.merge_cells('A'+str(row-1)+':'+mat_perc_delivered+str(row-1))
#             abm.row_dimensions[row].height = 37
#             abm.row_dimensions[row-1].height = 37
#             abm['A'+str(row-1)].font = Font(bold=True, sz=20, color=white)
#             abm['A'+str(row-1)].fill = PatternFill(
#                 fill_type='solid', start_color=black, end_color=black)
#             carbon_headers = ['A'+str(row), 'B'+str(row), 'C'+str(row), 'D'+str(row), 'E'+str(row), 'F'+str(row), 'G'+str(row), 'H'+str(row),
#                               'I'+str(row), 'J'+str(row), 'K'+str(row), 'L'+str(row), 'M'+str(row), 'N'+str(row)]
#             for cell in carbon_headers:
#                 abm[cell].font = Font(bold=True, color=white)
#                 abm[cell].fill = PatternFill(
#                     fill_type='solid', start_color=header_blue, end_color=header_blue)


# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

# SAVE AS NEW WORKBOOK
wb.save(filename='database/'+str(job_number)+'/JobSummary.xlsx')

print('Summary Complete!')
