import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, colors, Alignment, PatternFill, Border
from datetime import date, datetime
import json

job = sys.argv[1]

# SAVE JSON FILE
# file1 = open('database/'+job_number+'/job.json', 'a')
file1 = open('database/6973/job.json', 'a')
file1.truncate(0)
file1.write(json.dumps(job, indent=4))
file1.close()
