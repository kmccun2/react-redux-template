import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, colors, Alignment, PatternFill, Border
from datetime import date, datetime
import json

job_number = '7114'

# LOAD WORKBOOKS
nwd_wb = load_workbook(filename='database/Templates/WeldDataTemplate.xlsx')

# SET ACTIVATIONS
nwd = nwd_wb.active

print('')
print('Combing weld data files to a single file...')
print('')

# Iterate through workbooks in Weld Data folder
for each in os.listdir('database/'+job_number+'/Weld Data'):

    # Update printout
    print('Adding welds from ' + each)

    # Load weld data workbook and activate sheet
    wd_wb = load_workbook(filename='database/' +
                          str(job_number)+'/Weld Data/'+each)
    wd = wd_wb.active

    for col in range(1, wd.max_column+1):
        if wd[get_column_letter(col)+'1'].value == 'PIPELINE-REFERENCE':
            wd_pipeline = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'N_S_':
            wd_size = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'WELD-NO':
            wd_weldno = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'WELD-TYPE':
            wd_weldtype = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'PIPING-SPEC':
            wd_spec = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'ISO NO':
            wd_iso = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'SPOOL-DRAWING-SEQUENCE-NUMBER':
            wd_sketch = get_column_letter(col)
        elif wd[get_column_letter(col)+'1'].value == 'SPOOL-ID':
            wd_pm = get_column_letter(col)
    for row in range(2, wd.max_row+1):
        nwd['A'+str(nwd.max_row+1)] = wd[wd_pipeline+str(row)].value
        nwd['B'+str(nwd.max_row)] = wd[wd_size+str(row)].value
        nwd['C'+str(nwd.max_row)] = wd[wd_weldno+str(row)].value
        nwd['D'+str(nwd.max_row)] = wd[wd_weldtype+str(row)].value
        nwd['E'+str(nwd.max_row)] = wd[wd_spec+str(row)].value
        nwd['F'+str(nwd.max_row)] = wd[wd_iso+str(row)].value
        nwd['G'+str(nwd.max_row)] = wd[wd_sketch+str(row)].value
        nwd['H'+str(nwd.max_row)] = wd[wd_pm+str(row)].value

# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

# SAVE AS NEW WORKBOOK
nwd_wb.save(filename='database/'+str(job_number)+'/man_hours.xlsx')

print('Summary Complete!')
