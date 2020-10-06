from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE, get_column_letter
from openpyxl.styles import Font, Fill, NamedStyle, Border, Side, colors, Alignment, PatternFill, Border
from datetime import date, datetime
import json

# LOAD WORKBOOKS
jobnum = '7114'
bom_wb = load_workbook(filename='database/'+jobnum+'/bom_import.xlsx')

# ACTIVATE SHEETS
bom = bom_wb.active

size = 'D'
item = 'I'
desc = 'E'
sched = 'N'
bomclass = 'O'


bom[sched+'1'] = 'SCHEDULE'
bom[bomclass+'1'] = 'CLASS'

for row in range(2, bom.max_row+1):

    print_sched = 'None'
    print_class = None

    if 'X' in bom[size+str(row)].value.upper():
        itemsize = float(bom[size+str(row)].value.upper().split('X')[0].replace('1 1/2', '1.5').replace(
            '1 1/4', '1.25').replace('2 1/2', '2.5').replace('3/4', '.75').replace('1/2', '.5').replace('1/4', '.25'))
    else:
        itemsize = float(bom[size+str(row)].value.replace('1 1/2', '1.5').replace(
            '1 1/4', '1.25').replace('2 1/2', '2.5').replace('3/4', '.75').replace('1/2', '.5').replace('1/4', '.25'))
    newdesc = bom[desc+str(row)].value.replace('SCH ',
                                               'S-').replace('SCH', 'S-').replace(' WT', '').split(' ')

    if ' X ' in bom[desc+str(row)].value.replace('SCH ',
                                                 'S-').replace('SCH', 'S-').replace(' WT', ''):
        for each in newdesc:
            if each == 'X':
                for i in range(len(newdesc)):
                    if newdesc[i] == 'X':
                        print_sched = newdesc[i-1] + 'X' + newdesc[i+1]
    else:
        for each in newdesc:
            if 'S-' in each:
                print_sched = each.replace(' X ', 'x').replace(
                    ' ', '').replace('S-', '')

            elif 'CL' in each:
                print_class = each.replace('CL', '')

            elif each == 'XS':
                print_sched = 'XS'

            elif each == '10S':
                print_sched = '10S'

            elif each == 'XXS':
                print_sched = 'XXS'

            elif 'STD' in each:
                print_sched = 'STD'

            elif '3000#' in each or each == '3M':
                print_class = '3000'

            elif '150#' in each:
                print_class = '150'

            elif '800#' in each:
                print_class = '800'

            if bom[item+str(row)].value == 'SUPPORTS':
                print_sched = 'None'

            # Change sizes of 40 and 80
            if itemsize < 12 and print_sched == '40':
                print_sched = 'STD'
            if itemsize < 10 and print_sched == '80':
                print_sched = 'XS'

    # PRINT TO WORKBOOK
    bom[sched+str(row)] = print_sched
    bom[bomclass+str(row)] = print_class

# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

# SAVE NEW SHEET
bom_wb.save(filename='database/'+jobnum+'/bom_export.xlsx')

print('')
print('Complete!')
print('')
